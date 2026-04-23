#!/usr/bin/env python3
"""wowaudit.py — produce the Lua data file consumed by the Boble Loot addon.

Two modes, auto-selected:

  API fetch (default):
      py tools/wowaudit.py [--api-key ...]

      Pulls the team roster live from https://wowaudit.com. The API key
      can come from --api-key, the WOWAUDIT_API_KEY env var, or a `.env`
      file at the repo root (see .env.example).

  Manual convert:
      py tools/wowaudit.py --wowaudit path/to/export.csv --bis bis.json

      Reads a CSV / XLSX export from wowaudit's "Character data" sheet
      (required columns: character, mplus_dungeons, attendance,
      items_received, plus per-item sim_<itemID> columns) plus a JSON
      file mapping "Name-Realm" -> [BiS itemIDs].

In both modes the output is written to ``../Data/BobleLoot_Data.lua``
relative to this script (i.e. the addon's own Data folder), unless
``--out`` is given.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import functools
import json
import os
import sys
import time as _time_module
import urllib.error
import urllib.request
from pathlib import Path

# --------------------------------------------------------------------------
# .env loader (zero-dependency)
# --------------------------------------------------------------------------

def _load_dotenv() -> None:
    """Load KEY=VALUE pairs from the nearest ``.env`` into os.environ
    without overwriting existing variables."""
    for d in (Path(__file__).resolve().parent, *Path(__file__).resolve().parents):
        candidate = d / ".env"
        if not candidate.is_file():
            continue
        try:
            text = candidate.read_text(encoding="utf-8")
        except OSError:
            return
        for raw in text.splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
        return


_load_dotenv()

# --------------------------------------------------------------------------
# Retry / backoff helpers (item 4.4)
# --------------------------------------------------------------------------

def _sleep(seconds: float) -> None:
    """Thin wrapper around time.sleep for test-monkeypatching."""
    _time_module.sleep(seconds)


def retry_with_backoff(
    fn,
    delays: list[float],
    cache_key_fn,
    rate_limit_warn_threshold: int = 5,
):
    """Wrap ``fn`` with exponential backoff and a cached-fallback final stage.

    Args:
        fn: The callable to retry. Must accept ``(url_or_path, api_key)``
            and raise ``urllib.error.URLError`` / ``urllib.error.HTTPError``
            on failure.
        delays: Ordered list of sleep durations (seconds) between attempts.
            ``len(delays) + 1`` total attempts are made before the cache
            fallback is tried.
        cache_key_fn: Callable ``(url, api_key) -> str`` that returns the
            cache key to pass to ``_read_cache``. Pass ``None`` to skip
            the cache fallback (exception is propagated instead).
        rate_limit_warn_threshold: Log a warning when
            ``X-RateLimit-Remaining`` (if present in the response headers)
            is at or below this value. Default 5.

    Returns:
        A wrapped callable with the same signature as ``fn``.
    """
    @functools.wraps(fn)
    def wrapper(url_or_path, api_key):
        last_exc = None
        for attempt, delay in enumerate(
            [None] + delays  # attempt 0 = no prior delay
        ):
            if delay is not None:
                _sleep(delay)
            try:
                result = fn(url_or_path, api_key)
                return result
            except (urllib.error.URLError, urllib.error.HTTPError) as exc:
                last_exc = exc
                # Log rate-limit warnings immediately.
                remaining = None
                if hasattr(exc, "headers") and exc.headers:
                    remaining = exc.headers.get("X-RateLimit-Remaining")
                if remaining is not None:
                    try:
                        rem_int = int(remaining)
                        if rem_int <= rate_limit_warn_threshold:
                            print(
                                f"[WARN] X-RateLimit-Remaining={rem_int} — "
                                f"approaching WoWAudit rate limit.",
                                file=sys.stderr,
                            )
                    except (ValueError, TypeError):
                        pass
                # Continue to next retry.
        # All retries exhausted — try cache fallback.
        if cache_key_fn is not None:
            cached = _read_cache(cache_key_fn(url_or_path, api_key))
            if cached is not None:
                print(
                    f"[WARN] All retries failed for {url_or_path!r}; "
                    f"using cached response.",
                    file=sys.stderr,
                )
                return cached
        raise last_exc

    return wrapper


# --------------------------------------------------------------------------
# constants
# --------------------------------------------------------------------------

DEFAULT_OUT = Path(__file__).resolve().parent.parent / "Data" / "BobleLoot_Data.lua"
REQUIRED_COLS = {"character", "mplus_dungeons", "attendance"}

TIERS_DIR = Path(__file__).resolve().parent / "tiers"
REPO_ROOT  = Path(__file__).resolve().parent.parent

API_BASE = "https://wowaudit.com/v1"


def _load_tier_preset(tier_name: str) -> dict:
    """Load a tier preset JSON file from ``tools/tiers/``.

    The file is looked up as ``tools/tiers/<tier_name>.json`` with
    case-insensitive matching and hyphens normalised to lower case.

    Args:
        tier_name: e.g. ``"TWW-S3"`` or ``"tww-s3"``.

    Returns:
        Dict with any subset of keys: ``ilvlFloor``, ``mplusCap``,
        ``historyDays``, ``softFloor``, ``bisPath``.

    Raises:
        SystemExit: If no matching preset file is found.
    """
    normalised = tier_name.strip().lower()
    candidate  = TIERS_DIR / f"{normalised}.json"
    if not candidate.is_file():
        available = sorted(p.stem for p in TIERS_DIR.glob("*.json"))
        sys.exit(
            f"Tier preset '{tier_name}' not found. "
            f"Available: {', '.join(available) or '(none)'}. "
            f"Preset files live in tools/tiers/."
        )
    try:
        with candidate.open(encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        sys.exit(f"Failed to load tier preset '{tier_name}': {exc}")


# --------------------------------------------------------------------------
# Tier configuration — YAML (item 4.1, supersedes JSON presets from 2A)
# --------------------------------------------------------------------------

TIER_CONFIG_PATH = Path(__file__).resolve().parent / "tier-config.yaml"


def _load_tier_config(tier_name: str) -> dict:
    """Load a tier entry from ``tools/tier-config.yaml``.

    Args:
        tier_name: Case-insensitive tier key, e.g. ``"TWW-S3"`` or
            ``"tww-s3"``. Hyphens and underscores are both accepted.

    Returns:
        Dict with any subset of keys: ``ilvlFloor``, ``mplusCap``,
        ``historyDays``, ``softFloor``, ``bisPath``.  Missing keys
        in the YAML entry are absent from the returned dict (not ``None``),
        so callers can distinguish "not configured" from "explicitly null".

    Raises:
        SystemExit: If ``tier-config.yaml`` is missing, malformed, or
            the requested tier name does not appear under ``tiers:``.
    """
    try:
        import yaml
    except ImportError:
        sys.exit(
            "PyYAML is required for --tier. Install with: pip install pyyaml"
        )

    if not TIER_CONFIG_PATH.is_file():
        sys.exit(
            f"tools/tier-config.yaml not found at {TIER_CONFIG_PATH}. "
            "Create it or use the per-tier JSON presets in tools/tiers/."
        )

    try:
        doc = yaml.safe_load(TIER_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception as exc:  # yaml.YAMLError or OSError
        sys.exit(f"Failed to parse tools/tier-config.yaml: {exc}")

    tiers: dict = doc.get("tiers") or {}
    normalised = tier_name.strip().lower()
    if normalised not in tiers:
        available = sorted(tiers.keys())
        sys.exit(
            f"Tier '{tier_name}' not found in tools/tier-config.yaml. "
            f"Available tiers: {', '.join(available) or '(none)'}."
        )

    entry: dict = tiers[normalised] or {}
    return {k: v for k, v in entry.items() if v is not None or k == "bisPath"}


# --------------------------------------------------------------------------
# Character rename / realm transfer (item 4.5)
# --------------------------------------------------------------------------

def _apply_renames(
    rows: list[dict],
    bis: dict[str, list[int]],
    renames: dict[str, str],
) -> dict:
    """Apply a character-rename map to rows and BiS keys.

    Args:
        rows: Per-character row dicts. ``row["character"]`` is updated
            in-place for any matching old name.
        bis: BiS mapping ``{ "Name-Realm": [itemIDs] }``. Any key that
            appears as an old name in ``renames`` is replaced.
        renames: Mapping ``{ "Old-Realm": "New-Realm" }``. Keys starting
            with ``_`` are treated as metadata and ignored.

    Returns:
        Dict with keys ``"rows"`` (list[dict]) and ``"bis"`` (dict),
        both with renames applied.
    """
    # Filter out metadata keys (e.g. _comment, _example).
    effective: dict[str, str] = {
        old: new
        for old, new in renames.items()
        if not old.startswith("_") and isinstance(new, str)
    }

    if not effective:
        return {"rows": rows, "bis": bis}

    # Rename row character keys.
    for row in rows:
        old_name = row.get("character", "")
        if old_name in effective:
            row["character"] = effective[old_name]

    # Rename BiS keys.
    new_bis: dict[str, list[int]] = {}
    for key, item_ids in bis.items():
        new_key = effective.get(key, key)
        new_bis[new_key] = item_ids

    return {"rows": rows, "bis": new_bis}


# --------------------------------------------------------------------------
# BiS derivation from wishlist sim scores
# --------------------------------------------------------------------------

def _derive_bis_from_rows(
    rows: list[dict],
    threshold: float = 2.0,
) -> dict[str, list[int]]:
    """Derive a BiS mapping from sim scores already present in assembled rows.

    For each character and each ``sim_<itemID>`` column, the item is included
    in that character's BiS list when its score is **strictly greater than**
    ``threshold``.  Negative scores (downgrades) are never included.

    This removes the most significant manual maintenance burden from the data
    pipeline.  The derived BiS is per-character, not per-spec.

    Args:
        rows: Assembled character rows as returned by ``fetch_rows`` or
            ``_read_table``.  Each row may contain zero or more ``sim_<id>``
            keys whose values are numeric sim percentages.
        threshold: Minimum sim percentage (exclusive) for an item to be
            considered BiS.  Default is ``2.0`` (a 2% upgrade).

    Returns:
        Mapping of ``"Name-Realm"`` to a sorted list of BiS item IDs.
        Characters with no qualifying items are omitted from the mapping.
    """
    result: dict[str, list[int]] = {}
    for row in rows:
        name = (row.get("character") or "").strip()
        if not name:
            continue
        qualifying: list[int] = []
        for key, val in row.items():
            if not (key.startswith("sim_") and key[4:].isdigit()):
                continue
            score = _to_float(val, default=0.0)
            if score > threshold:
                qualifying.append(int(key[4:]))
        if qualifying:
            result[name] = sorted(qualifying)
    return result


# --------------------------------------------------------------------------
# BiS loader — file or versioned directory
# --------------------------------------------------------------------------

def load_bis(path: Path) -> dict[str, list[int]]:
    """Load a BiS mapping from a JSON file or a directory of JSON files.

    When ``path`` is a file, it is read directly and expected to be a JSON
    object of the form ``{ "Name-Realm": [itemID, ...] }``.

    When ``path`` is a directory, every ``*.json`` file in the directory
    tree (at any depth) is discovered and merged. This supports the versioned
    layout ``bis/<tier>/<class>-<spec>.json``. When the same character appears
    in multiple files, their item-ID lists are merged with deduplication.

    Non-``.json`` files (README, notes, etc.) are silently ignored.

    Args:
        path: A ``Path`` to either a JSON file or a directory.

    Returns:
        Mapping of ``"Name-Realm"`` to a deduplicated sorted list of BiS
        item IDs (as integers).

    Raises:
        SystemExit: If ``path`` does not exist.
        json.JSONDecodeError: If any discovered file contains invalid JSON.
        OSError: If a file cannot be opened.
    """
    if not path.exists():
        sys.exit(f"--bis path not found: {path}")

    if path.is_file():
        with path.open(encoding="utf-8") as f:
            raw: dict = json.load(f)
        return {
            k: [int(x) for x in v]
            for k, v in raw.items()
            if not k.startswith("_") and isinstance(v, list)
        }

    # Directory: walk recursively, collecting all .json files.
    merged: dict[str, set[int]] = {}
    for json_file in sorted(path.rglob("*.json")):
        with json_file.open(encoding="utf-8") as f:
            data: dict = json.load(f)
        for name, ids in data.items():
            # Skip metadata keys (e.g. "_comment") and non-list values.
            if name.startswith("_") or not isinstance(ids, list):
                continue
            if name not in merged:
                merged[name] = set()
            merged[name].update(int(x) for x in ids)

    return {name: sorted(ids) for name, ids in merged.items()}


# --------------------------------------------------------------------------
# CSV / XLSX reading
# --------------------------------------------------------------------------

def _read_table(path: Path) -> list[dict]:
    """Read a CSV or XLSX file into a list of dicts.

    For XLSX files, requires openpyxl (sys.exit with install hint if absent —
    this is an unrecoverable user-input error, not a network partial failure,
    so sys.exit is acceptable here per Batch 1A design decision).
    """
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            sys.exit("openpyxl is required to read XLSX inputs (pip install openpyxl)")
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {header[i]: row[i] for i in range(len(header)) if header[i]}
            for row in rows[1:]
            if any(c is not None for c in row)
        ]
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace("%", "").replace(",", ""))
    except ValueError:
        return default


def _to_int(v, default=0) -> int:
    return int(round(_to_float(v, default)))


def _lua_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')

# --------------------------------------------------------------------------
# Schema validation
# --------------------------------------------------------------------------

def _load_schema() -> dict | None:
    """Load tools/schemas/wowaudit_v1.json; return None if missing or invalid."""
    schema_path = Path(__file__).resolve().parent / "schemas" / "wowaudit_v1.json"
    if not schema_path.is_file():
        return None
    try:
        return json.loads(schema_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _validate_endpoint(
    data: object,
    endpoint_key: str,
    schema: dict | None,
    warnings: list[str],
) -> None:
    """Validate `data` against `schema['$defs'][endpoint_key]`.

    Appends a warning string to `warnings` on failure; never raises.
    Returns silently if jsonschema is not installed.
    """
    if schema is None:
        return
    try:
        import jsonschema  # type: ignore
    except ImportError:
        return
    sub = schema.get("$defs", {}).get(endpoint_key)
    if sub is None:
        return
    try:
        jsonschema.validate(data, sub)
    except jsonschema.ValidationError as exc:
        warnings.append(
            f"Schema validation failed for {endpoint_key!r}: {exc.message}"
        )

# --------------------------------------------------------------------------
# Cache helpers
# --------------------------------------------------------------------------

CACHE_DIR = Path(__file__).resolve().parent / ".cache"


def _cache_path(label: str) -> Path:
    """Return the cache file path for a given endpoint label."""
    # Sanitise label so it is safe as a filename on all platforms.
    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in label)
    return CACHE_DIR / f"{safe}.json"


def _write_cache(label: str, data: object) -> None:
    """Write `data` to the cache file for `label`. Silent on failure."""
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        _cache_path(label).write_text(json.dumps(data, indent=2), encoding="utf-8")
    except OSError:
        pass


def _read_cache(label: str) -> object | None:
    """Return cached data for `label`, or None if not found / unreadable."""
    p = _cache_path(label)
    if not p.is_file():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None

# --------------------------------------------------------------------------
# Lua emitter
# --------------------------------------------------------------------------

def build_lua(
    rows: list[dict],
    bis: dict[str, list[int]],
    sim_cap: float,
    mplus_cap: int,
    history_cap: int,
    team_url: str | None = None,
    fetch_warnings: list[str] | None = None,
    loot_min_ilvl: int = 0,
    history_days: int | None = None,
    tier_name: str | None = None,
    missing_wishlists: list[str] | None = None,
    renames: dict[str, str] | None = None,
    score_overrides: dict[int, float] | None = None,
) -> str:
    """Render BobleLoot_Data.lua from assembled rows.

    Args:
        rows: Per-character dicts with keys matching REQUIRED_COLS.
        bis: Mapping of "Name-Realm" to list of BiS item IDs.
        sim_cap: Maximum sim percentage (for Lua consumers).
        mplus_cap: Maximum M+ dungeons cap.
        history_cap: Maximum items-received cap.
        team_url: Optional WoWAudit team URL to embed.
        fetch_warnings: Optional list of warning strings from fetch_rows;
            emitted as Lua comments and a dataWarnings array.
        loot_min_ilvl: Minimum item level for loot history (from --tier preset
            or --loot-min-ilvl). Emitted when non-zero.
        history_days: Optional history window override (from --tier preset).
        tier_name: Optional tier preset name (e.g. "TWW-S3").
        missing_wishlists: Optional list of "Name-Realm" strings for characters
            present in the roster but absent from the /wishlists payload.
            Emitted as a ``missingWishlists`` array for the Lua side to surface.
        renames: Optional mapping ``{ "Old-Realm": "New-Realm" }`` emitted as
            a ``renames`` table for ``LootHistory:Apply`` to resolve stale keys
            (item 4.5). Keys starting with ``_`` are metadata and ignored.
        score_overrides: Optional mapping ``{ itemID: float }`` emitted as a
            ``scoreOverrides`` table for ``Scoring:Compute`` early-return
            (item 4.7). Non-numeric values are skipped silently.

    Returns:
        The Lua file contents as a string.

    Raises:
        ValueError: If rows is empty or missing required columns.
    """
    if not rows:
        raise ValueError("No rows to emit.")

    missing = REQUIRED_COLS - set(rows[0].keys())
    if missing:
        raise ValueError(f"Input is missing required columns: {sorted(missing)}")

    sim_cols_set: set[str] = set()
    for r in rows:
        for c in r.keys():
            if c.startswith("sim_") and c[4:].isdigit():
                sim_cols_set.add(c)
    sim_cols = sorted(sim_cols_set, key=lambda c: int(c[4:]))

    # Capture the generation moment once so ISO string and Unix timestamp match.
    now_utc   = dt.datetime.now(dt.timezone.utc)
    now_iso   = now_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
    now_epoch = int(now_utc.timestamp())
    out: list[str] = []

    # Warning comment block at the very top of the file.
    if fetch_warnings:
        for w in fetch_warnings:
            out.append(f"-- WARNING: {w}")
        out.append("")

    out.append("-- AUTO-GENERATED by tools/wowaudit.py - do not edit by hand.")
    out.append(f"-- Generated: {now_iso}")
    out.append("BobleLoot_Data = {")
    out.append(f'    generatedAt          = "{now_iso}",')
    # Unix integer form consumed by Lua (plan 1D freshness badge, item 1.6).
    # Emitted alongside generatedAt because Lua 5.1 has no standard date parser.
    out.append(f"    generatedAtTimestamp = {now_epoch},")
    if team_url:
        out.append(f'    teamUrl     = "{team_url}",')
    out.append(f"    simCap      = {sim_cap},")
    out.append(f"    mplusCap    = {mplus_cap},")
    out.append(f"    historyCap  = {history_cap},")
    if tier_name:
        out.append(f'    tierPreset  = "{_lua_escape(tier_name)}",')
    if loot_min_ilvl:
        out.append(f"    lootMinIlvl = {loot_min_ilvl},")
    if history_days is not None:
        out.append(f"    historyDays = {history_days},")

    # dataWarnings array — Lua side reads this to surface issues in the UI.
    if fetch_warnings:
        escaped = ", ".join(f'"{_lua_escape(w)}"' for w in fetch_warnings)
        out.append(f"    dataWarnings = {{ {escaped} }},")

    # missingWishlists array — characters in the roster with no wishlist data.
    if missing_wishlists:
        escaped_names = ", ".join(f'"{_lua_escape(n)}"' for n in missing_wishlists)
        out.append(f"    missingWishlists = {{ {escaped_names} }},")

    # Emit scoreOverrides table (item 4.7).
    # TODO(4B): include scoreOverrides in export bundle.
    effective_overrides: dict[int, float] = {}
    for raw_id, raw_val in (score_overrides or {}).items():
        try:
            iid = int(raw_id)
            fval = float(raw_val)
            effective_overrides[iid] = fval
        except (ValueError, TypeError):
            pass  # skip non-numeric entries silently

    if effective_overrides:
        out.append("    scoreOverrides = {")
        for item_id, score_val in sorted(effective_overrides.items()):
            out.append(f"        [{item_id}] = {score_val:.1f},")
        out.append("    },")

    # Emit renames table for LootHistory:Apply (item 4.5).
    effective_renames = {
        k: v for k, v in (renames or {}).items()
        if not k.startswith("_") and isinstance(v, str)
    }
    if effective_renames:
        out.append("    renames = {")
        for old_name, new_name in sorted(effective_renames.items()):
            out.append(
                f'        ["{_lua_escape(old_name)}"] = '
                f'"{_lua_escape(new_name)}",'
            )
        out.append("    },")

    out.append("    characters  = {")

    for row in rows:
        name = (row.get("character") or "").strip()
        if not name:
            continue
        attendance = _to_float(row.get("attendance"))
        mplus      = _to_int(row.get("mplus_dungeons"))

        out.append(f'        ["{_lua_escape(name)}"] = {{')
        out.append(f"            attendance    = {attendance},")
        out.append(f"            mplusDungeons = {mplus},")

        # mainspec and role — only emit if present (convert-mode CSVs won't
        # have these columns; that's fine, Scoring.lua treats absent = raider).
        mainspec_val = row.get("mainspec")
        role_val     = row.get("role", "raider")
        if mainspec_val:
            out.append(f'            mainspec      = "{_lua_escape(mainspec_val)}",')
        out.append(f'            role          = "{_lua_escape(role_val)}",')

        bis_ids = bis.get(name) or []
        if bis_ids:
            ids = ", ".join(f"[{int(i)}] = true" for i in bis_ids)
            out.append(f"            bis  = {{ {ids} }},")
        else:
            out.append("            bis  = {},")

        sim_pairs: list[str] = []
        sim_known_ids: list[int] = []
        for col in sim_cols:
            raw = row.get(col)
            if raw is None or raw == "":
                continue
            val = _to_float(raw, default=None)
            if val is None:
                continue
            item_id = int(col[4:])
            # simsKnown records every item the sim engine produced a result
            # for, even a 0% result. sims only carries the numeric value when
            # non-zero (file-size optimisation); Scoring.lua reads simsKnown
            # to tell "no data" from "data, value is 0".
            sim_known_ids.append(item_id)
            if val != 0.0:
                sim_pairs.append(f"[{item_id}] = {val}")
        if sim_pairs:
            out.append(f"            sims = {{ {', '.join(sim_pairs)} }},")
        else:
            out.append("            sims = {},")
        if sim_known_ids:
            known_pairs = ", ".join(f"[{i}] = true" for i in sim_known_ids)
            out.append(f"            simsKnown = {{ {known_pairs} }},")
        else:
            out.append("            simsKnown = {},")

        out.append("        },")

    out.append("    },")
    out.append("}")
    out.append("")
    return "\n".join(out)

# --------------------------------------------------------------------------
# Export bundle (roadmap 4.3)
# --------------------------------------------------------------------------

def export_bundle(
    rows: list[dict],
    bis: dict[str, list[int]],
    sim_cap: float,
    mplus_cap: int,
    history_cap: int,
    team_url: str | None = None,
    weights: dict | None = None,
) -> dict:
    """Build and return the portable export bundle as a Python dict.

    Args:
        rows:         Same row list used by build_lua().
        bis:          BiS mapping {name: [itemIDs]}.
        sim_cap:      Sim cap value used in this run.
        mplus_cap:    M+ cap value used in this run.
        history_cap:  History cap value used in this run.
        team_url:     Optional wowaudit team URL.
        weights:      Optional scoring weights dict (sim/bis/history/attendance/mplus).
                      If None, the BobleLoot defaults are used.

    Returns:
        A dict suitable for json.dumps() matching the bobleloot-export-v1 schema.
    """
    now_iso = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    default_weights = {
        "sim": 0.40, "bis": 0.20, "history": 0.15,
        "attendance": 0.15, "mplus": 0.10,
    }

    sim_cols_set: set[str] = set()
    for r in rows:
        for c in r.keys():
            if c.startswith("sim_") and c[4:].isdigit():
                sim_cols_set.add(c)

    characters: dict = {}
    for row in rows:
        name = (row.get("character") or "").strip()
        if not name:
            continue
        bis_ids = bis.get(name) or []
        sims: dict[int, float] = {}
        for col in sim_cols_set:
            val_raw = row.get(col)
            if val_raw not in (None, ""):
                val = _to_float(val_raw, default=None)
                if val is not None:
                    sims[int(col[4:])] = val
        char: dict = {
            "attendance":    _to_float(row.get("attendance")),
            "mplusDungeons": _to_int(row.get("mplus_dungeons")),
            "bis":           [int(i) for i in bis_ids],
            "sims":          sims,
        }
        if row.get("mainspec"):
            char["mainspec"] = str(row["mainspec"])
        if row.get("role"):
            char["role"] = str(row["role"])
        characters[name] = char

    bundle = {
        "schema":        "bobleloot-export-v1",
        "exportedAt":    now_iso,
        "generatedAt":   now_iso,
        "scoringConfig": {
            "simCap":     sim_cap,
            "mplusCap":   mplus_cap,
            "historyCap": history_cap,
            "weights":    weights or default_weights,
        },
        "characters":    characters,
    }
    if team_url:
        bundle["teamUrl"] = team_url
    return bundle


# --------------------------------------------------------------------------
# API fetch
# --------------------------------------------------------------------------

def http_get_json(path: str, api_key: str) -> object:
    """Fetch JSON from the WoWAudit API.

    Raises urllib.error.HTTPError, urllib.error.URLError, or OSError on
    failure. Never calls sys.exit — callers are responsible for handling
    exceptions and accumulating warnings.
    """
    url = API_BASE + path
    req = urllib.request.Request(
        url,
        headers={
            "Authorization": api_key,
            "Accept":        "application/json",
            "User-Agent":    "BobleLoot/0.1 (+wowaudit.py)",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


# Apply retry_with_backoff to http_get_json (item 4.4).
# Cache key is the path string (first argument), matching what _write_cache uses.
http_get_json = retry_with_backoff(
    http_get_json,
    delays=[5, 30],
    cache_key_fn=lambda path, _key: path,
)


def _full_name(name: str | None, realm: str | None) -> str | None:
    if not name:
        return None
    if realm:
        # Strip spaces from realm to match WoW's "Name-Realm" format.
        realm_clean = "".join(realm.split())
        return f"{name}-{realm_clean}"
    return name


def _best_wishlist_score(item: dict) -> float:
    """Pick the highest sim percentage across all specs/wishes for an item.

    wowaudit's API returns `score_by_spec[spec].percentage` and
    `wishes[*].percentage` as percent values already (e.g. 2.9 means a
    2.9% upgrade, 0.0744 means a 0.0744% sidegrade). We pass them
    through unchanged.
    """
    best = 0.0
    sbs = item.get("score_by_spec")
    if isinstance(sbs, dict):
        for spec_data in sbs.values():
            if isinstance(spec_data, dict):
                p = spec_data.get("percentage")
                if isinstance(p, (int, float)) and p > best:
                    best = p
    for w in item.get("wishes") or []:
        if isinstance(w, dict):
            p = w.get("percentage")
            if isinstance(p, (int, float)) and p > best:
                best = p
    return float(best)


def _mainspec_sim_score(item: dict, mainspec: str | None) -> float | None:
    """Return the sim percentage for the character's main spec only.

    Args:
        item: A wishlist item dict as returned by the WoWAudit API.
        mainspec: The spec name to match, e.g. ``"Holy"`` or ``"Protection"``.
            Case-insensitive prefix match is used so ``"holy"`` matches
            ``"Holy Paladin"`` if wowaudit ever returns a combined label.

    Returns:
        The percentage float for the matching spec, or ``None`` when the
        spec is not found in ``score_by_spec``.  Returns ``None`` (not 0.0)
        so callers can fall back to ``_best_wishlist_score`` rather than
        silently scoring zero.
    """
    if not mainspec:
        return None
    sbs = item.get("score_by_spec")
    if not isinstance(sbs, dict):
        return None
    target = mainspec.lower()
    for spec_key, spec_data in sbs.items():
        if not isinstance(spec_key, str):
            continue
        if spec_key.lower().startswith(target) or target.startswith(spec_key.lower()):
            if isinstance(spec_data, dict):
                p = spec_data.get("percentage")
                if isinstance(p, (int, float)):
                    return float(p)
    return None


# Determine role from WoWAudit member status.
# "raider" / "trial" / "bench" are first-class; anything else
# (social, unknown, absent) maps to "raider" as a safe default
# so scoring is never accidentally suppressed for a real raider
# whose API field uses an unfamiliar label.
_ROLE_MAP = {
    "trial": "trial",
    "bench": "bench",
}


def fetch_rows(
    api_key: str,
    dump_dir: Path | None,
    use_cache: bool = False,
    spec_aware: bool = True,
) -> tuple[list[dict], int, list[str], list[str]]:
    """Hit all wowaudit endpoints and merge into per-character row dicts.

    Returns:
        (rows, weeks_in_season, fetch_warnings, missing_wishlists)

    fetch_warnings is a list of human-readable warning strings for endpoints
    that failed or produced schema violations. The caller embeds these in the
    generated Lua file.

    missing_wishlists is a list of "Name-Realm" strings for roster characters
    that were absent from the /wishlists payload. These are characters whose
    sim data could not be populated and whose sims tables will be empty.
    """
    schema = _load_schema()
    fetch_warnings: list[str] = []

    def _fetch(path: str, label: str) -> object | None:
        """Fetch one endpoint; fall back to cache on error.

        Returns the parsed JSON, or None if both live fetch and cache fail.
        Appends to fetch_warnings on any failure.
        """
        if use_cache:
            cached = _read_cache(label)
            if cached is not None:
                return cached
            fetch_warnings.append(
                f"{label}: --use-cache requested but no cache file found; "
                "attempted live fetch."
            )

        try:
            data = http_get_json(path, api_key)
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")[:300]
            msg = f"{label}: HTTP {e.code} — {body}"
            fetch_warnings.append(msg)
            cached = _read_cache(label)
            if cached is not None:
                fetch_warnings.append(f"{label}: using cached fallback.")
                return cached
            return None
        except urllib.error.URLError as e:
            msg = f"{label}: network error — {e.reason}"
            fetch_warnings.append(msg)
            cached = _read_cache(label)
            if cached is not None:
                fetch_warnings.append(f"{label}: using cached fallback.")
                return cached
            return None
        except Exception as e:
            msg = f"{label}: unexpected error — {e}"
            fetch_warnings.append(msg)
            return None

        _write_cache(label, data)
        if dump_dir is not None:
            dump_dir.mkdir(parents=True, exist_ok=True)
            (dump_dir / f"{label}.json").write_text(
                json.dumps(data, indent=2), encoding="utf-8"
            )
        _validate_endpoint(data, f"{label}_response", schema, fetch_warnings)
        return data

    # --- period ---
    period_info = _fetch("/period", "period") or {}
    season      = (period_info or {}).get("current_season") or {}
    first_pid   = season.get("first_period_id")
    cur_pid     = (period_info or {}).get("current_period")
    season_start = season.get("start_date")

    # --- roster ---
    roster_raw = _fetch("/characters", "characters")
    if roster_raw is None:
        fetch_warnings.append("characters: endpoint failed and no cache — output will be empty.")
        roster = []
    elif not isinstance(roster_raw, list):
        fetch_warnings.append(
            f"characters: unexpected shape {type(roster_raw).__name__!r} — expected list."
        )
        roster = []
    else:
        roster = roster_raw

    # --- attendance ---
    att_path = "/attendance"
    if season_start:
        att_path += f"?start_date={season_start}"
    att_label = "attendance"
    attendance_payload = _fetch(att_path, att_label)
    att_chars = (
        attendance_payload.get("characters", [])
        if isinstance(attendance_payload, dict)
        else []
    )
    attendance_by_id = {
        c["id"]: c.get("attended_percentage", 0)
        for c in att_chars
        if isinstance(c, dict) and "id" in c
    }

    # --- historical M+ data ---
    dungeons_by_id: dict[int, int] = {}
    weeks_in_season = 0
    if isinstance(first_pid, int) and isinstance(cur_pid, int) and first_pid <= cur_pid:
        weeks_in_season = cur_pid - first_pid + 1
        for pid in range(first_pid, cur_pid + 1):
            label = f"historical_{pid}"
            payload = _fetch(f"/historical_data?period={pid}", label)
            for c in (
                (payload or {}).get("characters", [])
                if isinstance(payload, dict)
                else []
            ):
                cid  = c.get("id")
                done = (c.get("data") or {}).get("dungeons_done") or []
                if isinstance(cid, int) and isinstance(done, list):
                    dungeons_by_id[cid] = dungeons_by_id.get(cid, 0) + len(done)

    # Build mainspec lookup from roster so wishlist loop can use it.
    mainspec_by_id: dict[int, str | None] = {}
    for c in roster:
        if isinstance(c, dict):
            cid = c.get("id")
            if isinstance(cid, int):
                raw = c.get("main_spec") or ""
                mainspec_by_id[cid] = raw.strip() or None

    # --- wishlists ---
    wl_payload = _fetch("/wishlists", "wishlists")
    wl_chars   = (
        wl_payload.get("characters", []) if isinstance(wl_payload, dict) else []
    )
    sims_by_id: dict[int, dict[int, float]] = {}
    for c in wl_chars:
        if not isinstance(c, dict):
            continue
        cid = c.get("id")
        if not isinstance(cid, int):
            continue
        sims_by_id.setdefault(cid, {})
        char_mainspec = mainspec_by_id.get(cid)
        for inst in c.get("instances") or []:
            for diff in inst.get("difficulties") or []:
                wl = diff.get("wishlist") or {}
                for enc in wl.get("encounters") or []:
                    for item in enc.get("items") or []:
                        iid = item.get("id")
                        if not isinstance(iid, int):
                            continue
                        if spec_aware:
                            score = _mainspec_sim_score(item, char_mainspec)
                            if score is None:
                                score = _best_wishlist_score(item)
                        else:
                            score = _best_wishlist_score(item)
                        # First sighting wins; later sightings only overwrite
                        # when strictly greater. This ensures a 0.0 result is
                        # recorded (so simsKnown picks it up downstream)
                        # instead of being silently dropped by `score > prev`
                        # when prev defaulted to 0.0.
                        if iid not in sims_by_id[cid] or score > sims_by_id[cid][iid]:
                            sims_by_id[cid][iid] = score

    # --- cross-reference roster vs wishlists to detect missing characters ---
    # wl_char_ids is the set of character IDs present in the wishlists payload.
    # Any roster character whose ID is absent gets a warning and is added to
    # missing_wishlists so the Lua output can surface the gap to the raid leader.
    wl_char_ids: set[int] = set(sims_by_id.keys())
    missing_wishlists: list[str] = []
    if wl_payload is not None:
        # Only cross-reference when the endpoint succeeded; if wishlists failed
        # entirely the warning is already accumulated from _fetch().
        for c in roster:
            if not isinstance(c, dict):
                continue
            cid  = c.get("id")
            full = _full_name(c.get("name"), c.get("realm"))
            if not isinstance(cid, int) or not full:
                continue
            if cid not in wl_char_ids:
                missing_wishlists.append(full)
                fetch_warnings.append(
                    f"{full}: missing from /wishlists payload — sim data will be empty."
                )

    # --- assemble rows ---
    rows: list[dict] = []
    for c in roster:
        if not isinstance(c, dict):
            continue
        cid  = c.get("id")
        full = _full_name(c.get("name"), c.get("realm"))
        if not full:
            continue
        raw_status  = c.get("status") or ""
        row: dict = {
            "character":      full,
            "mplus_dungeons": dungeons_by_id.get(cid, 0),
            "attendance":     attendance_by_id.get(cid, 0),
            "role":           _ROLE_MAP.get(raw_status.lower(), "raider"),
            "mainspec":       (c.get("main_spec") or "").strip() or None,
        }
        for iid, score in (sims_by_id.get(cid) or {}).items():
            row[f"sim_{iid}"] = score
        rows.append(row)

    return rows, weeks_in_season, fetch_warnings, missing_wishlists


def fetch_team_url(api_key: str) -> str | None:
    try:
        team = http_get_json("/team", api_key)
        if isinstance(team, dict):
            return team.get("url")
    except Exception:
        pass
    return None


# --------------------------------------------------------------------------
# Run report
# --------------------------------------------------------------------------

def _parse_lua_names(lua_text: str) -> set[str]:
    """Extract character names from a BobleLoot_Data.lua file.

    Uses simple string scanning — not a full Lua parser. Matches lines of
    the form:  ["Name-Realm"] = {
    """
    import re
    return set(re.findall(r'\["([^"]+)"\]\s*=\s*\{', lua_text))


def _parse_lua_mplus_cap(lua_text: str) -> int | None:
    """Extract the mplusCap value from a BobleLoot_Data.lua file."""
    import re
    m = re.search(r'mplusCap\s*=\s*(\d+)', lua_text)
    return int(m.group(1)) if m else None


def _parse_lua_bis(lua_text: str) -> dict[str, set[int]]:
    """Extract BiS item IDs per character from a BobleLoot_Data.lua file.

    Returns {"Name-Realm": {itemID, ...}}.

    Uses a two-step approach to handle nested braces correctly: first find
    all character name positions, then find all bis = {...} blocks, and
    associate each bis block with the nearest preceding character name.
    """
    import re
    result: dict[str, set[int]] = {}
    char_positions = [
        (m.start(), m.group(1))
        for m in re.finditer(r'\["([^"]+)"\]\s*=\s*\{', lua_text)
    ]
    bis_blocks = [
        (m.start(), m.group(1))
        for m in re.finditer(r'bis\s*=\s*\{([^}]*)\}', lua_text)
    ]
    for bis_pos, bis_content in bis_blocks:
        preceding = [(pos, name) for pos, name in char_positions if pos < bis_pos]
        if preceding:
            _, name = max(preceding, key=lambda x: x[0])
            ids = set(int(x) for x in re.findall(r'\[(\d+)\]\s*=\s*true', bis_content))
            result[name] = ids
    return result


def _count_zero_sim_chars(rows: list[dict]) -> list[str]:
    """Return names of characters whose every sim column is 0 or absent."""
    zero_names: list[str] = []
    for row in rows:
        sim_vals = [
            v for k, v in row.items()
            if k.startswith("sim_") and k[4:].isdigit()
        ]
        if not sim_vals or all(_to_float(v) == 0.0 for v in sim_vals):
            zero_names.append(row.get("character", "?"))
    return zero_names


def _build_run_report(
    rows: list[dict],
    bis: dict[str, list[int]],
    mplus_cap: int,
    fetch_warnings: list[str],
    prev_lua_path: Path | None,
) -> str:
    """Build a human-readable run report string.

    Args:
        rows: The assembled character rows for this run.
        bis: BiS mapping used for this run.
        mplus_cap: The computed M+ cap for this run.
        fetch_warnings: Warnings accumulated during fetch.
        prev_lua_path: Path to the existing Lua file (for diffing); None if new.

    Returns:
        A multi-line string suitable for printing to stdout.
    """
    lines: list[str] = []
    lines.append("=" * 60)
    lines.append("BobleLoot run report")
    lines.append("=" * 60)

    # Characters.
    new_names = {(row.get("character") or "").strip() for row in rows if row.get("character")}
    lines.append(f"Characters this run : {len(new_names)}")

    # Diff against previous file.
    prev_names: set[str] = set()
    prev_mplus_cap: int | None = None
    prev_bis: dict[str, set[int]] = {}
    if prev_lua_path is not None and prev_lua_path.is_file():
        try:
            prev_text = prev_lua_path.read_text(encoding="utf-8")
            prev_names = _parse_lua_names(prev_text)
            prev_mplus_cap = _parse_lua_mplus_cap(prev_text)
            prev_bis = _parse_lua_bis(prev_text)
        except OSError:
            pass

    added   = sorted(new_names - prev_names)
    removed = sorted(prev_names - new_names)
    if added:
        lines.append(f"  Added   : {', '.join(added)}")
    if removed:
        lines.append(f"  Removed : {', '.join(removed)}")
    if not added and not removed and prev_names:
        lines.append("  Roster  : no change")

    # Zero-sim characters.
    zero_sims = _count_zero_sim_chars(rows)
    if zero_sims:
        lines.append(f"Zero sim data ({len(zero_sims)}) : {', '.join(zero_sims)}")
    else:
        lines.append("Zero sim data : none")

    # M+ cap.
    if prev_mplus_cap is not None and prev_mplus_cap != mplus_cap:
        lines.append(f"M+ cap : {prev_mplus_cap} -> {mplus_cap}")
    else:
        lines.append(f"M+ cap : {mplus_cap}")

    # BiS diff (summarised as total items changed).
    new_bis_sets = {name: set(ids) for name, ids in bis.items()}
    bis_changes: list[str] = []
    for name in sorted(new_bis_sets.keys() | prev_bis.keys()):
        old_set = prev_bis.get(name, set())
        new_set = new_bis_sets.get(name, set())
        if old_set != new_set:
            added_ids   = sorted(new_set - old_set)
            removed_ids = sorted(old_set - new_set)
            parts: list[str] = []
            if added_ids:
                parts.append(f"+{len(added_ids)} item(s)")
            if removed_ids:
                parts.append(f"-{len(removed_ids)} item(s)")
            bis_changes.append(f"  {name}: {', '.join(parts)}")
    if bis_changes:
        lines.append(f"BiS diff ({len(bis_changes)} character(s) changed):")
        lines.extend(bis_changes)
    else:
        lines.append("BiS diff : no change")

    # Fetch warnings.
    if fetch_warnings:
        lines.append(f"Warnings ({len(fetch_warnings)}):")
        for w in fetch_warnings:
            lines.append(f"  ! {w}")
    else:
        lines.append("Warnings : none")

    lines.append("=" * 60)
    return "\n".join(lines)

# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # Mode selection (auto): if --wowaudit is given, convert mode; else API.
    ap.add_argument("--wowaudit", type=Path, default=None,
                    help="CSV/XLSX export. Switches to convert mode (no API call).")
    ap.add_argument("--bis",      type=Path, default=None,
                    help="JSON file: { \"Name-Realm\": [itemID, ...] }. "
                         "Required in convert mode; optional in API mode.")
    # API mode
    ap.add_argument("--api-key", default=os.environ.get("WOWAUDIT_API_KEY"),
                    help="WoWAudit team API key (or WOWAUDIT_API_KEY env var, or .env file).")
    ap.add_argument("--dump-raw", type=Path, default=None,
                    help="Directory to also write raw API responses into (debugging).")
    ap.add_argument("--use-cache", action="store_true",
                    help="Replay the last successful API responses from "
                         "tools/.cache/ instead of hitting the network.")
    # Output
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help=f"Output .lua path (default: {DEFAULT_OUT}).")
    ap.add_argument("--sim-cap",     type=float, default=5.0)
    ap.add_argument("--mplus-cap",   type=int,   default=None,
                    help="Override M+ dungeons cap. Default: 10 per week of "
                         "the current season (so it grows over the season).")
    ap.add_argument("--mplus-cap-per-week", type=int, default=10,
                    help="Per-week increment for the auto M+ cap (default 10).")
    ap.add_argument("--history-cap", type=int,   default=5)
    ap.add_argument(
        "--tier",
        default=None,
        metavar="NAME",
        help=(
            "Apply a named tier preset from tools/tiers/<NAME>.json. "
            "Sets ilvlFloor, mplusCap, historyDays, softFloor, and optionally "
            "bisPath. Example: --tier TWW-S3. "
            "Explicit --mplus-cap / --history-cap / --loot-min-ilvl values "
            "always override the preset."
        ),
    )
    ap.add_argument(
        "--loot-min-ilvl",
        type=int,
        default=None,
        help=(
            "Minimum item level for loot history entries. "
            "Overrides the tier preset's ilvlFloor if both are specified."
        ),
    )
    ap.add_argument(
        "--no-spec-aware",
        action="store_true",
        default=False,
        help=(
            "Revert sim selection to max-across-all-specs (pre-2.1 behaviour). "
            "Default is spec-aware: only the character's main spec's sim is used."
        ),
    )
    ap.add_argument(
        "--bis-from-wishlist",
        action="store_true",
        default=False,
        help=(
            "Derive BiS membership from WoWAudit wishlist sim scores. "
            "Any item whose sim score exceeds --bis-threshold (default 2.0%%) "
            "is marked BiS for that character. Replaces --bis when both are given."
        ),
    )
    ap.add_argument(
        "--bis-threshold",
        type=float,
        default=2.0,
        help=(
            "Sim percentage threshold for --bis-from-wishlist (default 2.0). "
            "Items with a score strictly greater than this value are marked BiS."
        ),
    )
    ap.add_argument(
        "--renames",
        type=Path,
        default=None,
        metavar="FILE",
        help=(
            "Path to a renames.json sidecar. "
            "Default: tools/renames.json (sibling of wowaudit.py). "
            "Format: {'Old-Realm': 'New-Realm'}."
        ),
    )
    ap.add_argument(
        "--score-overrides",
        type=Path,
        default=None,
        metavar="FILE",
        help=(
            "Path to a score-overrides.json sidecar. "
            "Default: tools/score-overrides.json. "
            "Format: {'itemID': float}."
        ),
    )
    ap.add_argument(
        "--export",
        type=Path,
        default=None,
        metavar="PATH",
        help=(
            "After building, write a portable JSON bundle to PATH "
            "(no API key embedded). Use /bl importpaste in-game to load it."
        ),
    )
    args = ap.parse_args()

    # Apply tier preset (values are only used as defaults if the
    # corresponding explicit CLI argument was not provided).
    # 4.1: now reads from tier-config.yaml via _load_tier_config().
    tier_preset: dict = {}
    if args.tier is not None:
        tier_preset = _load_tier_config(args.tier)
        # Wire bisPath: if tier-config.yaml specifies a bisPath and --bis
        # was not explicitly provided, set args.bis automatically.
        preset_bis = tier_preset.get("bisPath")
        if preset_bis and args.bis is None:
            args.bis = REPO_ROOT / preset_bis

    def _preset(key: str, cli_val, default):
        """Return cli_val if it was explicitly set, else preset value, else default."""
        if cli_val is not None:
            return cli_val
        if key in tier_preset and tier_preset[key] is not None:
            return tier_preset[key]
        return default

    # Resolve loot min ilvl (used later in main for the run report / Lua header).
    loot_min_ilvl = _preset("ilvlFloor", args.loot_min_ilvl, 0)
    # Resolve history days override.
    history_days_override = _preset("historyDays", None, None)
    # Resolve soft floor (history cap) — used as new default for history_cap.
    soft_floor_override = _preset("softFloor", None, None)

    weeks_in_season = 1
    team_url = None
    fetch_warnings: list[str] = []
    missing_wishlists: list[str] = []
    spec_aware = not args.no_spec_aware
    if args.wowaudit is not None:
        # Convert mode.
        if args.bis is None:
            sys.exit("--bis is required when using --wowaudit.")
        rows = _read_table(args.wowaudit)
        bis = load_bis(args.bis)
    else:
        # API fetch mode.
        if not args.api_key:
            sys.exit(
                "No API key. Pass --api-key, set WOWAUDIT_API_KEY, "
                "or put it in a .env file. Alternatively pass --wowaudit "
                "to convert a manual export."
            )
        rows, weeks_in_season, fetch_warnings, missing_wishlists = fetch_rows(
            args.api_key, args.dump_raw, use_cache=args.use_cache,
            spec_aware=spec_aware,
        )
        team_url = fetch_team_url(args.api_key)
        if not rows and not fetch_warnings:
            sys.exit("No characters parsed from the API response.")
        if args.bis_from_wishlist:
            # Derive BiS from sim scores above the threshold.
            if args.bis is not None:
                fetch_warnings.append(
                    "--bis and --bis-from-wishlist both specified; "
                    "--bis-from-wishlist takes precedence."
                )
            bis = _derive_bis_from_rows(rows, threshold=args.bis_threshold)
        elif args.bis is not None:
            bis = load_bis(args.bis)
        else:
            # wowaudit's API doesn't expose a BiS flag; supply --bis to populate.
            bis = {}

    if args.mplus_cap is not None:
        mplus_cap = args.mplus_cap
    else:
        preset_mplus = tier_preset.get("mplusCap")
        if preset_mplus is not None:
            mplus_cap = int(preset_mplus)
        else:
            mplus_cap = max(args.mplus_cap_per_week,
                            args.mplus_cap_per_week * max(weeks_in_season, 1))

    # History cap: use soft_floor_override from preset if no explicit --history-cap provided.
    history_cap = args.history_cap
    if soft_floor_override is not None and args.history_cap == 5:
        # Only override the default (5) if the user didn't explicitly set it.
        history_cap = int(soft_floor_override)

    # Load rename sidecar (item 4.5).
    # --renames overrides the default sibling path (for CI with custom paths).
    renames_path = (
        args.renames if args.renames is not None
        else Path(__file__).resolve().parent / "renames.json"
    )
    renames: dict[str, str] = {}
    if renames_path.is_file():
        try:
            raw_renames = json.loads(renames_path.read_text(encoding="utf-8"))
            renames = {
                k: v for k, v in raw_renames.items()
                if isinstance(k, str) and isinstance(v, str)
                and not k.startswith("_")
            }
        except (json.JSONDecodeError, OSError) as exc:
            print(f"[WARN] Failed to load {renames_path}: {exc}", file=sys.stderr)

    # Apply renames to rows and BiS keys before emission.
    renamed = _apply_renames(rows, bis, renames)
    rows = renamed["rows"]
    bis  = renamed["bis"]

    # Load score-overrides sidecar (item 4.7).
    # --score-overrides overrides the default sibling path.
    overrides_path = (
        args.score_overrides if args.score_overrides is not None
        else Path(__file__).resolve().parent / "score-overrides.json"
    )
    score_overrides: dict[int, float] = {}
    if overrides_path.is_file():
        try:
            raw_overrides = json.loads(overrides_path.read_text(encoding="utf-8"))
            for k, v in raw_overrides.items():
                if k.startswith("_"):
                    continue
                try:
                    score_overrides[int(k)] = float(v)
                except (ValueError, TypeError):
                    print(
                        f"[WARN] score-overrides.json: skipping invalid entry "
                        f"{k!r}={v!r}",
                        file=sys.stderr,
                    )
        except (json.JSONDecodeError, OSError) as exc:
            print(f"[WARN] Failed to load {overrides_path}: {exc}",
                  file=sys.stderr)

    try:
        lua = build_lua(
            rows, bis, args.sim_cap, mplus_cap, history_cap,
            team_url, fetch_warnings if not args.wowaudit else None,
            loot_min_ilvl=loot_min_ilvl,
            history_days=history_days_override,
            tier_name=args.tier,
            missing_wishlists=missing_wishlists if not args.wowaudit else None,
            renames=renames,
            score_overrides=score_overrides,
        )
    except ValueError as exc:
        sys.exit(str(exc))
    report = _build_run_report(
        rows, bis, mplus_cap, fetch_warnings if not args.wowaudit else [],
        args.out if args.out.is_file() else None,
    )
    print(report)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(lua, encoding="utf-8")
    print(f"Wrote {args.out}: {len(rows)} characters, "
          f"{sum(len(v) for v in bis.values())} BiS entries. "
          f"M+ cap = {mplus_cap} ({weeks_in_season} week(s) into season).")

    # Export bundle (roadmap 4.3): --export writes a portable JSON bundle.
    if args.export is not None:
        bundle = export_bundle(
            rows, bis,
            sim_cap=args.sim_cap,
            mplus_cap=mplus_cap,
            history_cap=history_cap,
            team_url=team_url,
        )
        args.export.parent.mkdir(parents=True, exist_ok=True)
        args.export.write_text(
            json.dumps(bundle, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"Exported bundle to {args.export}: "
              f"{len(bundle['characters'])} characters.")


if __name__ == "__main__":
    main()
